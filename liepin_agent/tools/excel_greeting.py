"""Excel-driven batch greeting workflow."""

from __future__ import annotations

import json
import inspect
import logging
import random
import time
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from threading import Event
from typing import Callable, Dict, List, Optional
from urllib.parse import urlsplit, urlunsplit

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


@dataclass
class ExcelGreetingCandidate:
    row_index: int
    name: str
    profile_url: str
    tier: str
    is_gold_collar: bool
    greeting_status: str = ""
    contact_text: str = ""
    current_company: str = ""
    current_title: str = ""


class ExcelGreetingService:
    STATUS_SUCCESS = "已发送"
    STATUS_ALREADY = "已打过"
    STATUS_FAILED = "失败"
    STATUS_SKIPPED = "已跳过"
    STATUS_DRY_RUN = "待发送"

    PERMANENT_STATUSES = {"already_greeted", "skipped", "dry_run"}

    def __init__(self, liepin_tool):
        self.liepin_tool = liepin_tool
        self._stop_event = Event()

    def request_stop(self) -> None:
        self._stop_event.set()

    @property
    def is_stopped(self) -> bool:
        return self._stop_event.is_set()

    @classmethod
    def load_greetable_candidates(cls, excel_path: str | Path, gold_only: bool = True) -> List[ExcelGreetingCandidate]:
        workbook = load_workbook(excel_path)
        try:
            sheet = workbook["候选人"] if "候选人" in workbook.sheetnames else workbook.active
            headers = cls._header_map(sheet)
            required = ["姓名", "匹配档位", "金领", "简历链接", "打招呼状态"]
            missing = [header for header in required if header not in headers]
            if missing:
                raise ValueError("Excel 缺少必要列：{}".format("、".join(missing)))
            result: List[ExcelGreetingCandidate] = []
            for row_index in range(2, sheet.max_row + 1):
                tier = cls._cell_text(sheet, row_index, headers, "匹配档位").upper()
                if tier not in {"A", "B"}:
                    continue
                if gold_only and not cls._is_yes(cls._cell_text(sheet, row_index, headers, "金领")):
                    continue
                greeting_status = cls._cell_text(sheet, row_index, headers, "打招呼状态")
                if greeting_status in {"已发送", "已打过", "success", "already_greeted"}:
                    continue
                contact_text = " ".join(
                    cls._cell_text(sheet, row_index, headers, header)
                    for header in ("联系方式", "打招呼消息", "打招呼错误")
                    if header in headers
                )
                profile_url = cls.normalize_profile_url(cls._cell_text(sheet, row_index, headers, "简历链接"))
                if not profile_url:
                    continue
                result.append(
                    ExcelGreetingCandidate(
                        row_index=row_index,
                        name=cls._cell_text(sheet, row_index, headers, "姓名") or "候选人",
                        profile_url=profile_url,
                        tier=tier,
                        is_gold_collar=cls._is_yes(cls._cell_text(sheet, row_index, headers, "金领")),
                        greeting_status=greeting_status,
                        contact_text=contact_text,
                        current_company=cls._cell_text(sheet, row_index, headers, "公司"),
                        current_title=cls._cell_text(sheet, row_index, headers, "职位"),
                    )
                )
            return result
        finally:
            workbook.close()

    def greet_from_excel(
        self,
        excel_path: str | Path,
        message_template: str = "",
        delay_min: float = 2.0,
        delay_max: float = 5.0,
        dry_run: bool = False,
        verify_gold_on_page: bool = True,
        request_resume: bool = False,
        gold_only: bool = True,
        max_retries: int = 1,
        max_candidates: int = 0,
        progress_callback: Optional[Callable[[int, int, str], None]] = None,
    ) -> List[Dict[str, object]]:
        candidates = self.load_greetable_candidates(excel_path, gold_only=gold_only)
        candidates.sort(key=lambda c: (c.tier != "A", c.row_index))
        if max_candidates > 0:
            candidates = candidates[:max_candidates]
        results: List[Dict[str, object]] = []
        pending_writes: List[Dict[str, object]] = []
        total = len(candidates)
        batch_write_interval = min(10, max(1, total // 5)) if total > 0 else 1
        for index, candidate in enumerate(candidates, start=1):
            if self._stop_event.is_set():
                break
            if progress_callback:
                progress_callback(index, total, candidate.name)
            if dry_run:
                results.append(
                    {
                        "row_index": candidate.row_index,
                        "candidate_name": candidate.name,
                        "profile_url": candidate.profile_url,
                        "status": "dry_run",
                        "message": "预览：未发送打招呼",
                        "error": "",
                    }
                )
                continue
            if index > 1:
                time.sleep(random.uniform(delay_min, delay_max))
            if self._stop_event.is_set():
                break
            payload = {
                "id": str(candidate.row_index),
                "name": candidate.name,
                "profile_url": candidate.profile_url,
                "current_company": candidate.current_company,
                "current_title": candidate.current_title,
                "is_gold_collar": candidate.is_gold_collar,
                "skip_gold_check": not verify_gold_on_page,
            }
            response = self._greet_with_retry(payload, message_template, request_resume, max_retries)
            status = str(response.get("status") or "failed")
            message = str(response.get("message") or "")
            error = str(response.get("error") or "")
            request_resume_status = str(response.get("request_resume_status") or "")
            result_item = {
                "row_index": candidate.row_index,
                "candidate_name": candidate.name,
                "status": status,
                "message": message,
                "error": error,
                "request_resume_status": request_resume_status,
            }
            results.append(result_item)
            pending_writes.append(result_item)
            if len(pending_writes) >= batch_write_interval:
                try:
                    self._write_batch_results(excel_path, pending_writes)
                    pending_writes = []
                except Exception as exc:
                    logger.warning("periodic batch write failed: %s", exc)
        if pending_writes:
            try:
                self._write_batch_results(excel_path, pending_writes)
            except Exception as exc:
                logger.warning("final batch write failed: %s", exc)
        try:
            self._write_batch_results(excel_path, results)
        except Exception as exc:
            logger.warning("full batch write results failed: %s", exc)
        return results

    def _greet_with_retry(
        self,
        payload: Dict[str, object],
        message_template: str,
        request_resume: bool,
        max_retries: int,
    ) -> Dict[str, str]:
        last_response: Dict[str, str] = {"status": "failed", "message": "", "error": "未知错误"}
        for attempt in range(1 + max(0, max_retries)):
            if self._stop_event.is_set():
                return {"status": "failed", "message": "", "error": "用户取消"}
            try:
                greet_candidate = self.liepin_tool.greet_candidate
                kwargs = {"message_template": message_template}
                if self._accepts_keyword(greet_candidate, "request_resume"):
                    kwargs["request_resume"] = request_resume
                response = greet_candidate(payload, **kwargs)
            except Exception as exc:
                response = {"status": "failed", "message": "", "error": str(exc)}
            last_response = response
            status = str(response.get("status") or "failed")
            if status != "failed" or status in self.PERMANENT_STATUSES:
                return response
            if attempt < max_retries:
                logger.warning(
                    "greet retry %d/%d for %s: %s",
                    attempt + 1, max_retries,
                    payload.get("name") or payload.get("id"),
                    response.get("error") or "",
                )
                time.sleep(random.uniform(2.0, 4.0))
        return last_response

    @staticmethod
    def _accepts_keyword(callback: Callable[..., object], keyword: str) -> bool:
        """Return whether a callable accepts a keyword without invoking it."""
        try:
            parameters = inspect.signature(callback).parameters
        except (TypeError, ValueError):
            # Some extension callables do not expose a signature. Preserve the
            # modern API in that uncommon case instead of disabling features.
            return True
        return keyword in parameters or any(
            parameter.kind == inspect.Parameter.VAR_KEYWORD
            for parameter in parameters.values()
        )

    @classmethod
    def write_greeting_result(
        cls,
        excel_path: str | Path,
        row_index: int,
        status: str,
        message: str = "",
        error: str = "",
        request_resume_status: str = "",
    ) -> None:
        if status == "dry_run":
            return
        import time as _time
        last_exc = None
        for attempt in range(3):
            try:
                workbook = load_workbook(excel_path)
                try:
                    sheet = workbook["候选人"] if "候选人" in workbook.sheetnames else workbook.active
                    headers = cls._header_map(sheet)
                    cls._ensure_header(sheet, headers, "打招呼状态")
                    cls._ensure_header(sheet, headers, "打招呼消息")
                    cls._ensure_header(sheet, headers, "打招呼错误")
                    cls._ensure_header(sheet, headers, "索要简历状态")
                    labels = {
                        "success": cls.STATUS_SUCCESS,
                        "already_greeted": cls.STATUS_ALREADY,
                        "skipped": cls.STATUS_SKIPPED,
                        "failed": cls.STATUS_FAILED,
                        "dry_run": cls.STATUS_DRY_RUN,
                    }
                    sheet.cell(row=row_index, column=headers["打招呼状态"]).value = labels.get(status, status)
                    sheet.cell(row=row_index, column=headers["打招呼消息"]).value = message or ""
                    sheet.cell(row=row_index, column=headers["打招呼错误"]).value = error or ""
                    sheet.cell(row=row_index, column=headers["索要简历状态"]).value = request_resume_status or ""
                    workbook.save(excel_path)
                finally:
                    workbook.close()
                return
            except PermissionError as exc:
                last_exc = exc
                if attempt < 2:
                    _time.sleep(1.5)
                continue
            except Exception:
                raise
        raise last_exc

    @classmethod
    def _write_batch_results(cls, excel_path: str | Path, results: List[Dict[str, object]]) -> None:
        """将一批结果统一写入 Excel，用于补救中间因文件占用而失败的单条写入。"""
        writable_results = [
            item for item in results if str(item.get("status") or "") != "dry_run"
        ]
        if not writable_results:
            return
        import time as _time
        for attempt in range(3):
            try:
                workbook = load_workbook(excel_path)
                try:
                    sheet = workbook["候选人"] if "候选人" in workbook.sheetnames else workbook.active
                    headers = cls._header_map(sheet)
                    cls._ensure_header(sheet, headers, "打招呼状态")
                    cls._ensure_header(sheet, headers, "打招呼消息")
                    cls._ensure_header(sheet, headers, "打招呼错误")
                    cls._ensure_header(sheet, headers, "索要简历状态")
                    labels = {
                        "success": cls.STATUS_SUCCESS,
                        "already_greeted": cls.STATUS_ALREADY,
                        "skipped": cls.STATUS_SKIPPED,
                        "failed": cls.STATUS_FAILED,
                        "dry_run": cls.STATUS_DRY_RUN,
                    }
                    for item in writable_results:
                        row_index = int(item.get("row_index") or 0)
                        if row_index <= 0:
                            continue
                        status = str(item.get("status") or "failed")
                        message = str(item.get("message") or "")
                        error = str(item.get("error") or "")
                        request_resume_status = str(item.get("request_resume_status") or "")
                        sheet.cell(row=row_index, column=headers["打招呼状态"]).value = labels.get(status, status)
                        sheet.cell(row=row_index, column=headers["打招呼消息"]).value = message or ""
                        sheet.cell(row=row_index, column=headers["打招呼错误"]).value = error or ""
                        sheet.cell(row=row_index, column=headers["索要简历状态"]).value = request_resume_status or ""
                    workbook.save(excel_path)
                finally:
                    workbook.close()
                return
            except PermissionError:
                if attempt < 2:
                    _time.sleep(1.5)
                continue
            except Exception:
                raise

    @staticmethod
    def generate_summary(results: List[Dict[str, object]], cancelled: bool = False) -> str:
        total = len(results)
        success = sum(1 for item in results if item.get("status") == "success")
        already = sum(1 for item in results if item.get("status") == "already_greeted")
        failed = sum(1 for item in results if item.get("status") == "failed")
        skipped = sum(1 for item in results if item.get("status") == "skipped")
        dry_run = sum(1 for item in results if item.get("status") == "dry_run")
        resume_success = sum(
            1 for item in results
            if item.get("request_resume_status") == "已发送索要简历"
        )
        lines = [
            "批量打招呼{}：共处理 {} 位候选人".format("（已取消）" if cancelled else "完成", total),
            "成功：{} 人".format(success),
            "已打过：{} 人".format(already),
            "跳过：{} 人".format(skipped),
            "失败：{} 人".format(failed),
        ]
        if resume_success:
            lines.append("已索要简历：{} 人".format(resume_success))
        if dry_run:
            lines.append("待发送：{} 人（dry-run 预览，未实际发送）".format(dry_run))
            names = "、".join(str(item.get("candidate_name") or "候选人") for item in results[:20])
            if len(results) > 20:
                names += " 等"
            lines.append("预览名单：{}".format(names or "-"))
        for item in results:
            if item.get("status") == "failed":
                lines.append("{}：{}".format(item.get("candidate_name") or "候选人", item.get("error") or "失败"))
        return "\n".join(lines)

    @staticmethod
    def normalize_profile_url(url: str) -> str:
        value = str(url or "").strip()
        if not value:
            return ""
        if value.startswith("/") and not value.startswith("//"):
            value = "https://h.liepin.com" + value
        elif value.startswith("//"):
            value = "https:" + value
        try:
            parsed = urlsplit(value)
        except ValueError:
            return ""
        if parsed.scheme not in {"http", "https"}:
            return ""
        host = (parsed.hostname or "").lower()
        if host != "liepin.com" and not host.endswith(".liepin.com"):
            return ""
        path = (parsed.path or "").lower()
        query = (parsed.query or "").lower()
        if "showresumedetail" not in path + "?" + query and "/resume/" not in path:
            return ""
        return urlunsplit(("https", parsed.netloc, parsed.path, parsed.query, parsed.fragment))

    @staticmethod
    def _header_map(sheet) -> Dict[str, int]:
        return {
            str(cell.value or "").strip(): index
            for index, cell in enumerate(sheet[1], start=1)
            if str(cell.value or "").strip()
        }

    @staticmethod
    def _ensure_header(sheet, headers: Dict[str, int], header: str) -> None:
        if header in headers:
            return
        column = sheet.max_column + 1
        sheet.cell(row=1, column=column).value = header
        headers[header] = column

    @staticmethod
    def _cell_text(sheet, row_index: int, headers: Dict[str, int], header: str) -> str:
        column = headers.get(header)
        if not column:
            return ""
        return str(sheet.cell(row=row_index, column=column).value or "").strip()

    @staticmethod
    def _is_yes(value: str) -> bool:
        return str(value or "").strip().lower() in {"是", "yes", "true", "1", "金领"}


class GreetingQuotaTracker:
    def __init__(self, workspace_root: str | Path):
        self._path = Path(workspace_root) / ".greeting_quota.json"

    def _load(self) -> Dict[str, object]:
        try:
            return json.loads(self._path.read_text(encoding="utf-8"))
        except Exception:
            return {}

    def _save(self, data: Dict[str, object]) -> None:
        try:
            self._path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
        except Exception:
            pass

    def today_count(self) -> int:
        data = self._load()
        if data.get("date") != str(date.today()):
            return 0
        return int(data.get("count") or 0)

    def increment(self, n: int = 1) -> int:
        data = self._load()
        today = str(date.today())
        if data.get("date") != today:
            data = {"date": today, "count": 0}
        data["count"] = int(data.get("count") or 0) + n
        self._save(data)
        return int(data["count"])
