"""Excel export from SQLite candidate data."""

from __future__ import annotations

import re
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional
from xml.sax.saxutils import escape as xml_escape

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

from ..storage.sqlite_store import SQLiteStore, from_json, now_text


class ExportService:
    _JOB_KEYWORDS = (
        "工程师",
        "经理",
        "总监",
        "主管",
        "专员",
        "顾问",
        "设计师",
        "开发",
        "运营",
        "产品经理",
        "销售",
        "会计",
        "人事",
        "行政",
        "财务",
        "采购",
        "物流",
        "分析师",
        "架构师",
        "测试",
        "运维",
        "市场",
        "品牌",
        "助理",
        "客服",
        "技术支持",
        "项目管理",
        "生产",
        "质量",
        "工艺",
        "制造",
        "设备",
        "机械",
        "电气",
        "自动化",
        "材料",
        "化工",
    )
    _COMPANY_MARKERS = (
        "有限公司",
        "有限责任公司",
        "股份公司",
        "公司",
        "集团",
        "研究院",
        "研究所",
        "事务所",
        "中心",
    )
    _INVALID_CITY_WORDS = (
        "工作",
        "经验",
        "求职",
        "期望",
        "求职期望",
        "不限",
        "统招",
        "全日制",
        "MBA/EMBA",
        "EMBA",
        "MBA",
    )
    _COMPANY_TITLE_SEPARATORS = (" · ", "·", "｜", "|")

    HEADERS = [
        "姓名",
        "公司",
        "职位",
        "城市",
        "年限",
        "学历",
        "金领",
        "来源",
        "卡片判断",
        "详情状态",
        "匹配档位",
        "证据分",
        "打招呼状态",
        "打招呼消息",
        "打招呼错误",
        "匹配摘要",
        "风险",
        "命中证据",
        "缺口/未知",
        "待确认问题",
        "置信度",
        "基准版本",
        "简历链接",
        "候选人ID",
        "原始卡片摘要",
    ]

    OVERVIEW_HEADERS = [
        "结论",
        "档位",
        "证据分",
        "姓名",
        "公司",
        "职位",
        "城市",
        "年限",
        "学历",
        "金领",
        "卡片判断",
        "详情",
        "打招呼",
        "匹配摘要",
        "主要风险",
        "缺口/未知",
        "待确认问题",
        "简历链接",
        "候选人档案",
        "候选人ID",
    ]

    def __init__(self, store: SQLiteStore, export_dir: str | Path):
        self.store = store
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)
        self.last_candidate_reports_dir: Optional[Path] = None

    def export_session(self, session_id: str) -> Path:
        session = self.store.get_session(session_id) or {}
        candidates = self._candidate_export_rows(session_id)
        base_filename = "{}_{}".format(
            self._safe_filename(str(session.get("title") or "候选人")),
            now_text().replace(":", "").replace(" ", "_"),
        )
        filename = "{}.xlsx".format(base_filename)
        path = self.export_dir / filename
        reports_dir = self.export_dir / "{}_候选人档案".format(base_filename)
        self._write_candidate_reports(reports_dir, session, candidates)
        self.last_candidate_reports_dir = reports_dir

        workbook = Workbook()
        workbook.active.title = "推荐总览"
        self._write_overview_sheet(workbook.active, candidates)
        self._write_overview_sheet(
            workbook.create_sheet("合格A_B"),
            [item for item in candidates if item["is_qualified"]],
        )
        self._write_overview_sheet(
            workbook.create_sheet("待复核_未匹配"),
            [item for item in candidates if item["is_unmatched"]],
        )
        self._write_overview_sheet(
            workbook.create_sheet("不合格C"),
            [item for item in candidates if item["is_rejected"]],
        )
        self._write_candidate_sheet(workbook.create_sheet("候选人"), candidates)
        self._write_criteria_sheet(workbook, session_id)
        self._write_metrics_sheet(workbook, session_id)
        self._write_diagnostics_sheet(workbook, session_id)
        workbook.save(path)
        workbook.close()
        return path

    def _candidate_export_rows(self, session_id: str) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        for item in self.store.list_candidates(session_id):
            candidate_id = str(item.get("id") or "")
            detail = self.store.get_candidate_detail(candidate_id) or {}
            sources = self.store.list_candidate_sources(candidate_id)
            fields = self._display_candidate_fields(item)
            profile_url = self._candidate_profile_url(item, detail)
            evidence = [
                evidence_item
                for evidence_item in (item.get("matched_evidence") or [])
                if isinstance(evidence_item, dict)
            ]
            tier = str(item.get("match_tier") or "").upper()
            row = {
                "item": item,
                "detail": detail,
                "fields": fields,
                "sources": sources,
                "profile_url": profile_url,
                "evidence": evidence,
                "source_text": self._source_text(sources),
                "evidence_text": self._evidence_text(evidence),
                "missing_text": "\n".join(item.get("missing_or_unclear") or []),
                "questions_text": "\n".join(item.get("questions_to_verify") or []),
                "conclusion": self._candidate_conclusion(item),
                "tier": tier,
                "is_qualified": tier in {"A", "B"},
                "is_rejected": tier in {"C", "D"},
                "is_unmatched": tier not in {"A", "B", "C", "D"},
                "report_path": None,
            }
            rows.append(row)
        return sorted(rows, key=self._export_sort_key)

    def _write_overview_sheet(
        self, sheet, candidates: List[Dict[str, Any]]
    ) -> None:
        sheet.append(self.OVERVIEW_HEADERS)
        for row in candidates:
            item = row["item"]
            fields = row["fields"]
            sheet.append(
                [
                    row["conclusion"],
                    row["tier"],
                    item.get("match_score") or 0,
                    fields.get("name") or "",
                    fields.get("current_company") or "",
                    fields.get("current_title") or "",
                    fields.get("city") or "",
                    fields.get("work_years") or "",
                    fields.get("education") or "",
                    "是" if int(item.get("is_gold_collar") or 0) == 1 else "否",
                    self._card_decision_label(item.get("card_decision") or ""),
                    item.get("detail_capture_status") or "",
                    self._greeting_status_label(item.get("greeting_status") or ""),
                    item.get("match_summary") or "",
                    item.get("match_risks") or "",
                    row["missing_text"],
                    row["questions_text"],
                    row["profile_url"],
                    self._report_display_name(row),
                    item.get("id") or "",
                ]
            )
            self._set_row_links(sheet, sheet.max_row, row, self.OVERVIEW_HEADERS)
        self._style_sheet(
            sheet,
            self.OVERVIEW_HEADERS,
            long_headers={"匹配摘要", "主要风险", "缺口/未知", "待确认问题"},
            tab_color="2563EB",
        )

    def _write_candidate_sheet(
        self, sheet, candidates: List[Dict[str, Any]]
    ) -> None:
        sheet.append(self.HEADERS)
        for row in candidates:
            item = row["item"]
            fields = row["fields"]
            sheet.append(
                [
                    fields.get("name") or "",
                    fields.get("current_company") or "",
                    fields.get("current_title") or "",
                    fields.get("city") or "",
                    fields.get("work_years") or "",
                    fields.get("education") or "",
                    "是" if int(item.get("is_gold_collar") or 0) == 1 else "否",
                    row["source_text"],
                    self._card_decision_label(item.get("card_decision") or ""),
                    item.get("detail_capture_status") or "",
                    item.get("match_tier") or "",
                    item.get("match_score") or 0,
                    self._greeting_status_label(item.get("greeting_status") or ""),
                    item.get("greeting_message") or "",
                    item.get("greeting_error") or "",
                    item.get("match_summary") or "",
                    item.get("match_risks") or "",
                    row["evidence_text"],
                    row["missing_text"],
                    row["questions_text"],
                    item.get("confidence") or "",
                    item.get("criteria_version_id") or "",
                    row["profile_url"],
                    item.get("id") or "",
                    fields.get("summary_text") or "",
                ]
            )
            self._set_row_links(sheet, sheet.max_row, row, self.HEADERS)
        self._style_sheet(
            sheet,
            self.HEADERS,
            long_headers={
                "来源",
                "打招呼消息",
                "打招呼错误",
                "匹配摘要",
                "风险",
                "命中证据",
                "缺口/未知",
                "待确认问题",
                "原始卡片摘要",
            },
            tab_color="64748B",
        )

    def _set_row_links(
        self, sheet, row_index: int, row: Dict[str, Any], headers: List[str]
    ) -> None:
        if row.get("profile_url") and "简历链接" in headers:
            link_column = headers.index("简历链接") + 1
            link_cell = sheet.cell(row=row_index, column=link_column)
            link_cell.hyperlink = row["profile_url"]
            link_cell.style = "Hyperlink"
        report_path = row.get("report_path")
        if report_path and "候选人档案" in headers:
            report_column = headers.index("候选人档案") + 1
            report_cell = sheet.cell(row=row_index, column=report_column)
            report_cell.hyperlink = self._file_hyperlink(report_path)
            report_cell.style = "Hyperlink"

    def _style_sheet(
        self,
        sheet,
        headers: List[str],
        long_headers: Optional[set[str]] = None,
        tab_color: str = "3B82F6",
    ) -> None:
        long_headers = long_headers or set()
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_properties.tabColor = tab_color
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=tab_color)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column, header in enumerate(headers, start=1):
            if header in long_headers:
                width = 42
            elif header in {"简历链接"}:
                width = 46
            elif header in {"候选人档案"}:
                width = 30
            elif header in {"候选人ID"}:
                width = 34
            else:
                width = 16
            sheet.column_dimensions[get_column_letter(column)].width = width
        sheet.row_dimensions[1].height = 24

    def _write_candidate_reports(
        self,
        reports_dir: Path,
        session: Dict[str, Any],
        candidates: List[Dict[str, Any]],
    ) -> None:
        reports_dir.mkdir(parents=True, exist_ok=True)
        for index, row in enumerate(candidates, start=1):
            report_path = reports_dir / self._candidate_report_filename(index, row)
            self._write_candidate_docx(report_path, session, row)
            row["report_path"] = report_path

    def _candidate_report_filename(self, index: int, row: Dict[str, Any]) -> str:
        fields = row["fields"]
        item = row["item"]
        name = fields.get("name") or "候选人"
        title = fields.get("current_title") or item.get("match_tier") or "档案"
        tier = row["tier"] or "未匹配"
        filename = "{:02d}_{}_{}_{}.docx".format(index, tier, name, title)
        return self._safe_filename(filename[:-5]) + ".docx"

    def _write_candidate_docx(
        self, path: Path, session: Dict[str, Any], row: Dict[str, Any]
    ) -> None:
        item = row["item"]
        detail = row["detail"]
        fields = row["fields"]
        lines: List[tuple[str, str]] = []

        def heading(text: str) -> None:
            lines.append(("heading", text))

        def paragraph(label: str, value: object = "") -> None:
            text = str(value or "").strip()
            if label and text:
                lines.append(("paragraph", "{}：{}".format(label, text)))
            elif label:
                lines.append(("paragraph", label))
            elif text:
                lines.append(("paragraph", text))

        title = "{} / {}".format(
            fields.get("name") or "候选人", fields.get("current_title") or "职位未知"
        )
        lines.append(("title", title))
        paragraph("任务", session.get("title") or "")
        paragraph("结论", "{} {}".format(row["conclusion"], row["tier"]).strip())
        paragraph(
            "基本信息",
            "{} | {} | {} | {} | 金领：{}".format(
                fields.get("current_company") or "公司未知",
                fields.get("city") or "城市未知",
                fields.get("work_years") or "年限未知",
                fields.get("education") or "学历未知",
                "是" if int(item.get("is_gold_collar") or 0) == 1 else "否",
            ),
        )
        paragraph("候选人链接", row["profile_url"])
        paragraph("候选人ID", item.get("id") or "")

        heading("匹配结论")
        paragraph("匹配档位", row["tier"] or "未匹配")
        paragraph("置信度", item.get("confidence") or "")
        paragraph("摘要", item.get("match_summary") or fields.get("summary_text") or "")
        paragraph("风险", item.get("match_risks") or "")

        heading("命中证据")
        if row["evidence"]:
            for index, evidence in enumerate(row["evidence"], start=1):
                paragraph(
                    "{}.".format(index),
                    "[{}] {} - {} ({})".format(
                        self._evidence_source_label(evidence),
                        evidence.get("criterion") or "",
                        evidence.get("evidence") or "",
                        evidence.get("strength") or "",
                    ),
                )
        else:
            paragraph("暂无命中证据")

        heading("缺口与待确认")
        paragraph("缺口/未知", row["missing_text"] or "暂无")
        paragraph("待确认问题", row["questions_text"] or "暂无")

        heading("来源")
        paragraph("", row["source_text"] or "暂无来源记录")

        heading("简历内容")
        paragraph("详情状态", item.get("detail_capture_status") or "")
        paragraph("简历摘要", detail.get("resume_summary") or "")
        paragraph("简历正文", detail.get("resume_text") or "尚未抓取简历详情。")

        heading("打招呼")
        paragraph("状态", self._greeting_status_label(item.get("greeting_status") or ""))
        paragraph("消息", item.get("greeting_message") or "")
        paragraph("错误", item.get("greeting_error") or "")

        self._write_docx(path, lines)

    def _display_candidate_fields(self, item: Dict[str, Any]) -> Dict[str, str]:
        fields = {
            "name": str(item.get("name") or ""),
            "current_company": str(item.get("current_company") or ""),
            "current_title": str(item.get("current_title") or ""),
            "city": str(item.get("city") or ""),
            "work_years": str(item.get("work_years") or ""),
            "education": str(item.get("education") or ""),
            "summary_text": str(item.get("summary_text") or ""),
        }
        parsed = self._parse_summary_fields(fields["summary_text"])
        if parsed.get("name") and not fields["name"]:
            fields["name"] = parsed["name"]
        split_company, split_title = self._split_company_title(
            fields["current_company"]
        )
        if split_company and split_title:
            if fields["current_title"] and split_company in self._COMPANY_MARKERS:
                fields["current_company"] = fields["current_title"] + split_company
                fields["current_title"] = split_title
            elif not fields["current_title"] or "·" in fields["current_company"]:
                fields["current_company"] = split_company
                fields["current_title"] = split_title
        split_company, split_title = self._split_company_title(fields["current_title"])
        if split_company and split_title and not fields["current_company"]:
            fields["current_company"] = split_company
            fields["current_title"] = split_title
        if parsed.get("current_company") and parsed.get("current_title"):
            fields["current_company"] = parsed["current_company"]
            fields["current_title"] = parsed["current_title"]
        for key in ("city", "work_years", "education"):
            if not parsed.get(key):
                continue
            if key == "city":
                if not self._looks_like_city(fields["city"]):
                    fields["city"] = parsed[key]
            elif not fields[key]:
                fields[key] = parsed[key]
        return fields

    def _candidate_profile_url(
        self, item: Dict[str, Any], detail: Optional[Dict[str, Any]] = None
    ) -> str:
        detail = detail if detail is not None else {}
        if not detail:
            candidate_id = str(item.get("id") or "")
            if candidate_id:
                detail = self.store.get_candidate_detail(candidate_id) or {}
        raw_payload = from_json(detail.get("raw_payload_json"), {}) or {}
        profile_url = self._find_url(raw_payload)
        if profile_url:
            return profile_url
        return str(item.get("profile_url") or "")

    @staticmethod
    def _source_text(sources: List[Dict[str, Any]]) -> str:
        return "\n".join(
            "{} / {} / 排名{}".format(
                source.get("query") or "",
                source.get("search_hypothesis_type") or "",
                source.get("result_index") or 0,
            )
            for source in sources
        )

    @staticmethod
    def _evidence_source_label(item: Dict[str, Any]) -> str:
        if item.get("source_type") == "inferred":
            return "推断"
        grounding_status = item.get("grounding_status")
        if grounding_status == "exact":
            return "原文证据"
        if grounding_status == "model_summary":
            return "模型概括"
        return "匹配证据"

    @staticmethod
    def _evidence_text(evidence: List[Dict[str, Any]]) -> str:
        return "\n".join(
            "[{}] {}: {}".format(
                ExportService._evidence_source_label(evidence_item),
                evidence_item.get("criterion") or "",
                evidence_item.get("evidence") or "",
            )
            for evidence_item in evidence
        )

    @classmethod
    def _candidate_conclusion(cls, item: Dict[str, Any]) -> str:
        tier = str(item.get("match_tier") or "").upper()
        if tier == "A":
            return "强推荐"
        if tier == "B":
            return "可推荐"
        if tier in {"C", "D"}:
            return "不推荐"
        if item.get("detail_capture_status"):
            return "待复核"
        return "未抓详情"

    @classmethod
    def _export_sort_key(cls, row: Dict[str, Any]) -> tuple[int, int, str]:
        tier = row["tier"]
        tier_order = {"A": 0, "B": 1, "C": 3}.get(tier, 2)
        score = int(row["item"].get("match_score") or 0)
        name = str(row["fields"].get("name") or "")
        return tier_order, -score, name

    @staticmethod
    def _report_display_name(row: Dict[str, Any]) -> str:
        report_path = row.get("report_path")
        return report_path.name if isinstance(report_path, Path) else ""

    @staticmethod
    def _file_hyperlink(path: Path) -> str:
        try:
            return path.resolve().as_uri()
        except ValueError:
            return str(path)

    @classmethod
    def _write_docx(cls, path: Path, lines: List[tuple[str, str]]) -> None:
        body = "\n".join(cls._docx_paragraph(style, text) for style, text in lines)
        document_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:body>
    {body}
    <w:sectPr>
      <w:pgSz w:w="11906" w:h="16838"/>
      <w:pgMar w:top="1440" w:right="1440" w:bottom="1440" w:left="1440" w:header="708" w:footer="708" w:gutter="0"/>
    </w:sectPr>
  </w:body>
</w:document>
""".format(body=body)
        path.parent.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
            archive.writestr(
                "[Content_Types].xml",
                """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>
""",
            )
            archive.writestr(
                "_rels/.rels",
                """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>
""",
            )
            archive.writestr("word/document.xml", document_xml)

    @staticmethod
    def _docx_paragraph(style: str, text: str) -> str:
        size = {"title": "32", "heading": "26"}.get(style, "22")
        bold = "<w:b/>" if style in {"title", "heading"} else ""
        spacing = "240" if style in {"title", "heading"} else "120"
        escaped = xml_escape(str(text or ""))
        return """
<w:p>
  <w:pPr><w:spacing w:after="{spacing}"/></w:pPr>
  <w:r>
    <w:rPr>{bold}<w:sz w:val="{size}"/></w:rPr>
    <w:t xml:space="preserve">{text}</w:t>
  </w:r>
</w:p>""".format(
            spacing=spacing,
            bold=bold,
            size=size,
            text=escaped,
        )

    @classmethod
    def _parse_summary_fields(cls, summary_text: str) -> Dict[str, str]:
        lines = [
            line.strip()
            for line in str(summary_text or "").splitlines()
            if line.strip()
        ]
        parsed = {
            "name": lines[0] if lines else "",
            "current_company": "",
            "current_title": "",
            "city": "",
            "work_years": "",
            "education": "",
        }
        full_text = " ".join(lines)
        work_match = re.search(r"(?:工作)?(\d+年(?:经验)?)", full_text)
        if work_match:
            parsed["work_years"] = work_match.group(1)
        education_match = re.search(
            r"(MBA/EMBA|EMBA|MBA|本科|硕士|博士|大专|中专|高中|初中)",
            full_text,
        )
        if education_match:
            parsed["education"] = education_match.group(1)
        for line in reversed(lines[1:]):
            company, title = cls._split_company_title(line)
            if company:
                parsed["current_company"] = company
                parsed["current_title"] = title
                break
        for line in lines[1:]:
            city = cls._strip_personal_line(line)
            if cls._looks_like_city(city):
                parsed["city"] = city
                break
        return parsed

    @classmethod
    def _split_company_title(cls, line: str) -> tuple[str, str]:
        value = (line or "").strip()
        if not value or "求职期望" in value:
            return "", ""
        for separator in cls._COMPANY_TITLE_SEPARATORS:
            if separator not in value:
                continue
            parts = [part.strip() for part in value.split(separator) if part.strip()]
            if len(parts) < 2:
                continue
            company = separator.join(parts[:-1]).strip()
            title = parts[-1].strip()
            if len(company) < 2 or len(title) < 2:
                continue
            if any(keyword in title for keyword in cls._JOB_KEYWORDS) or any(
                marker in company for marker in cls._COMPANY_MARKERS
            ):
                return company, title
        return "", ""

    @classmethod
    def _strip_personal_line(cls, line: str) -> str:
        text = (line or "").strip()
        text = re.sub(r"\d+岁", "", text)
        text = re.sub(r"(?:工作)?\d+年(?:经验)?", "", text)
        text = re.sub(
            r"(MBA/EMBA|EMBA|MBA|本科|硕士|博士|大专|中专|高中|初中)", "", text
        )
        text = re.sub(r"\d+k(?:-\d+k)?", "", text)
        text = re.sub(r"^(男|女)\s*", "", text)
        for tag in ("男", "女", "已婚", "未婚", "共青团员", "党员", "群众"):
            text = text.replace(tag, "")
        return text.replace(" ", "").strip()

    @classmethod
    def _looks_like_city(cls, value: str) -> bool:
        text = (value or "").strip()
        if (
            not text
            or text in cls._INVALID_CITY_WORDS
            or not (2 <= len(text) <= 12)
            or re.search(r"\d", text)
        ):
            return False
        if any(word in text for word in cls._INVALID_CITY_WORDS):
            return False
        if any(keyword in text for keyword in cls._JOB_KEYWORDS):
            return False
        if any(marker in text for marker in cls._COMPANY_MARKERS):
            return False
        if re.search(
            r"(MBA/EMBA|EMBA|MBA|本科|硕士|博士|大专|中专|高中|初中|\d+岁)",
            text,
        ):
            return False
        return True

    @classmethod
    def _find_url(cls, value: Any) -> str:
        if isinstance(value, str):
            if value.startswith(("http://", "https://")):
                return value
            nested = from_json(value, None)
            return cls._find_url(nested) if nested is not None else ""
        if isinstance(value, dict):
            for key in ("profile_url", "resume_url", "detail_url", "url", "href"):
                url = value.get(key)
                if isinstance(url, str) and url.startswith(("http://", "https://")):
                    return url
            for child in value.values():
                url = cls._find_url(child)
                if url:
                    return url
        if isinstance(value, list):
            for child in value:
                url = cls._find_url(child)
                if url:
                    return url
        return ""

    def _write_criteria_sheet(self, workbook: Workbook, session_id: str) -> None:
        criteria = self.store.get_latest_criteria_version(session_id, "confirmed") or {}
        sheet = workbook.create_sheet("寻访基准")
        sheet.append(["版本", criteria.get("version") or ""])
        sheet.append(["关键词", criteria.get("keywords_text") or ""])
        sheet.append(["岗位要求", criteria.get("requirements_text") or ""])
        sheet.column_dimensions["A"].width = 16
        sheet.column_dimensions["B"].width = 80

    def _write_metrics_sheet(self, workbook: Workbook, session_id: str) -> None:
        sheet = workbook.create_sheet("效率总结")
        metrics = self.store.session_efficiency_metrics(session_id)
        for key, value in metrics.items():
            sheet.append([key, value])
        sheet.append([])
        sheet.append(["搜索假设", "描述", "轮次", "读卡片", "去重候选", "抓详情", "A/B", "噪音", "重复"])
        for item in self.store.search_hypothesis_metrics(session_id):
            sheet.append(
                [
                    item.get("search_hypothesis_type") or "",
                    item.get("search_hypothesis_text") or "",
                    item.get("round_count") or 0,
                    item.get("raw_count") or 0,
                    item.get("unique_count") or 0,
                    item.get("detail_fetch_count") or 0,
                    item.get("ab_count") or 0,
                    item.get("noise_count") or 0,
                    item.get("duplicate_count") or 0,
                ]
            )
        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 56

    def _write_diagnostics_sheet(self, workbook: Workbook, session_id: str) -> None:
        sheet = workbook.create_sheet("运行诊断")
        summary = self.store.session_diagnostic_summary(session_id)
        sheet.append(["诊断项", "值"])
        for flag in summary.get("diagnostic_flags") or []:
            sheet.append(["建议", flag])
        sheet.append(["待回写匹配数", summary.get("pending_match_count") or 0])
        sheet.append(["错误事件数", summary.get("error_count") or 0])
        sheet.append([])
        sheet.append(["轮次状态", "数量"])
        for key, value in (summary.get("round_status_counts") or {}).items():
            sheet.append([key, value])
        sheet.append([])
        sheet.append(["卡片判断", "数量"])
        for key, value in (summary.get("card_decision_counts") or {}).items():
            sheet.append([key, value])
        sheet.append([])
        sheet.append(["详情状态", "数量"])
        for key, value in (summary.get("detail_status_counts") or {}).items():
            sheet.append([key, value])
        sheet.append([])
        sheet.append(["匹配状态", "数量"])
        for key, value in (summary.get("match_status_counts") or {}).items():
            sheet.append([key, value])
        sheet.append([])
        sheet.append(["匹配档位", "数量"])
        for key, value in (summary.get("tier_counts") or {}).items():
            sheet.append([key or "未定档", value])
        sheet.append([])
        sheet.append(["最近错误", "消息", "时间"])
        for item in summary.get("recent_errors") or []:
            sheet.append(
                [
                    item.get("title") or "",
                    item.get("message") or "",
                    item.get("created_at") or "",
                ]
            )
        self._style_sheet(
            sheet,
            ["诊断项", "值", "时间"],
            long_headers={"值", "消息"},
            tab_color="DC2626",
        )

    @staticmethod
    def _safe_filename(value: str) -> str:
        value = re.sub(r'[\\/:*?"<>|\x00-\x1f]+', "_", value)
        return value.strip(" .")[:60].rstrip(" .") or "候选人"

    @staticmethod
    def _card_decision_label(value: object) -> str:
        return {
            "fetch": "值得抓详情",
            "maybe": "信息不足",
            "noise": "明显噪音",
        }.get(str(value or ""), "信息不足")

    @staticmethod
    def _greeting_status_label(value: object) -> str:
        return {
            "success": "已发送",
            "already_greeted": "已打过",
            "skipped": "已跳过",
            "failed": "失败",
            "pending": "待处理",
        }.get(str(value or ""), "")
