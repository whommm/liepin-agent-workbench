import pytest
from openpyxl import Workbook, load_workbook

from liepin_agent.tools.excel_greeting import ExcelGreetingService
from liepin_agent.tools.greeting_text import GreetingTextGenerationService
from liepin_agent.tools.real_liepin import RealLiepinTool
from liepin_agent.domain.recommendation import (
    HIGH_POTENTIAL_VERIFY,
    INFORMATION_INSUFFICIENT,
    PRIORITY_CONTACT,
    TRANSFERABLE_EXPLORE,
)


def _make_excel(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "候选人"
    sheet.append([
        "姓名",
        "公司",
        "职位",
        "金领",
        "建议状态",
        "打招呼状态",
        "简历链接",
    ])
    sheet.append(["优先候选人", "能源公司", "销售总监", "是", "优先沟通", "", "https://h.liepin.com/resume/showresumedetail/?res_id=a"])
    sheet.append(["高潜候选人", "能源公司", "销售经理", "是", "高潜待确认", "", "/resume/showresumedetail/?res_id=b"])
    sheet.append(["迁移候选人", "能源公司", "销售", "是", "可迁移探索", "", "https://h.liepin.com/resume/showresumedetail/?res_id=c"])
    sheet.append(["非金领", "能源公司", "销售", "否", "优先沟通", "", "https://h.liepin.com/resume/showresumedetail/?res_id=d"])
    sheet.append(["已打过", "能源公司", "销售", "是", "优先沟通", "已发送", "https://h.liepin.com/resume/showresumedetail/?res_id=e"])
    sheet.append(["外站", "能源公司", "销售", "是", "优先沟通", "", "https://example.com/resume/showresumedetail/?res_id=f"])
    workbook.save(path)
    workbook.close()


def _make_recommendation_excel(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "候选人"
    sheet.append([
        "姓名",
        "公司",
        "职位",
        "金领",
        "建议状态",
        "打招呼状态",
        "简历链接",
    ])
    rows = [
        ("优先", "优先沟通", "priority"),
        ("高潜", "high_potential_verify", "potential"),
        ("迁移", "可迁移探索", "transfer"),
        ("不足", "信息不足", "insufficient"),
        ("不匹配", "明确不匹配", "mismatch"),
    ]
    for name, state, resume_id in rows:
        sheet.append([
            name,
            "能源公司",
            "销售",
            "是",
            state,
            "",
            "https://h.liepin.com/resume/showresumedetail/?res_id={}".format(resume_id),
        ])
    workbook.save(path)
    workbook.close()


def test_excel_greeting_loads_default_recommendation_states(tmp_path):
    path = tmp_path / "candidates.xlsx"
    _make_excel(path)

    candidates = ExcelGreetingService.load_greetable_candidates(path)

    assert [item.name for item in candidates] == ["优先候选人", "高潜候选人"]
    assert candidates[1].profile_url == "https://h.liepin.com/resume/showresumedetail/?res_id=b"


def test_excel_greeting_rejects_retired_tier_only_workbook(tmp_path):
    path = tmp_path / "legacy.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["姓名", "金领", "匹配档位", "打招呼状态", "简历链接"])
    sheet.append(["旧候选人", "是", "A", "", "https://h.liepin.com/resume/showresumedetail/?res_id=old"])
    workbook.save(path)
    workbook.close()

    with pytest.raises(ValueError, match="建议状态"):
        ExcelGreetingService.load_greetable_candidates(path)


def test_excel_greeting_filters_by_selected_recommendation_states(tmp_path):
    path = tmp_path / "recommendations.xlsx"
    _make_recommendation_excel(path)

    default_candidates = ExcelGreetingService.load_greetable_candidates(path)
    assert [item.name for item in default_candidates] == ["优先", "高潜"]
    assert [item.recommendation_state for item in default_candidates] == [
        PRIORITY_CONTACT,
        HIGH_POTENTIAL_VERIFY,
    ]

    selected = ExcelGreetingService.load_greetable_candidates(
        path,
        recommendation_states=[TRANSFERABLE_EXPLORE, INFORMATION_INSUFFICIENT],
    )
    assert [item.name for item in selected] == ["迁移", "不足"]


def test_excel_greeting_never_bulk_loads_explicit_mismatch(tmp_path):
    path = tmp_path / "recommendations.xlsx"
    _make_recommendation_excel(path)

    candidates = ExcelGreetingService.load_greetable_candidates(
        path,
        recommendation_states=["explicit_mismatch"],
    )

    assert candidates == []


def test_excel_greeting_dry_run_does_not_send_or_write(tmp_path):
    path = tmp_path / "candidates.xlsx"
    _make_excel(path)
    original_bytes = path.read_bytes()

    class Tool:
        def greet_candidate(self, candidate, message_template=""):
            raise AssertionError("dry-run should not send greetings")

    results = ExcelGreetingService(Tool()).greet_from_excel(path, dry_run=True)

    assert [item["status"] for item in results] == ["dry_run", "dry_run"]
    assert path.read_bytes() == original_bytes
    workbook = load_workbook(path)
    sheet = workbook["候选人"]
    headers = {cell.value: index for index, cell in enumerate(sheet[1], start=1)}
    assert sheet.cell(row=2, column=headers["打招呼状态"]).value in {None, ""}
    workbook.close()


def test_excel_greeting_passes_request_resume_when_tool_supports_it(tmp_path):
    path = tmp_path / "candidates.xlsx"
    _make_excel(path)
    request_resume_values = []

    class Tool:
        def greet_candidate(
            self, candidate, message_template="", request_resume=False
        ):
            request_resume_values.append(request_resume)
            return {"status": "success", "message": "ok", "error": ""}

    ExcelGreetingService(Tool()).greet_from_excel(
        path,
        delay_min=0,
        delay_max=0,
        request_resume=True,
    )

    assert request_resume_values == [True, True]


def test_excel_greeting_rechecks_gold_by_default(tmp_path):
    path = tmp_path / "candidates.xlsx"
    _make_excel(path)
    payloads = []

    class Tool:
        def greet_candidate(self, candidate, message_template=""):
            payloads.append(candidate)
            return {"status": "success", "message": "ok", "error": ""}

    ExcelGreetingService(Tool()).greet_from_excel(path, delay_min=0, delay_max=0)

    assert [payload["skip_gold_check"] for payload in payloads] == [False, False]


def test_excel_greeting_can_explicitly_trust_excel_gold(tmp_path):
    path = tmp_path / "candidates.xlsx"
    _make_excel(path)
    payloads = []

    class Tool:
        def greet_candidate(self, candidate, message_template=""):
            payloads.append(candidate)
            return {"status": "success", "message": "ok", "error": ""}

    ExcelGreetingService(Tool()).greet_from_excel(
        path,
        delay_min=0,
        delay_max=0,
        verify_gold_on_page=False,
    )

    assert [payload["skip_gold_check"] for payload in payloads] == [True, True]


def test_liepin_profile_url_validation():
    assert RealLiepinTool._ensure_absolute_url("/resume/showresumedetail/?res_id=1") == "https://h.liepin.com/resume/showresumedetail/?res_id=1"
    assert RealLiepinTool._ensure_absolute_url("https://h.liepin.com/resume/showresumedetail/?res_id=1") == "https://h.liepin.com/resume/showresumedetail/?res_id=1"
    assert RealLiepinTool._ensure_absolute_url("https://example.com/resume/showresumedetail/?res_id=1") == ""
    assert RealLiepinTool._ensure_absolute_url("javascript:alert(1)") == ""
    assert RealLiepinTool._ensure_absolute_url("https://h.liepin.com/") == ""


def test_excel_greeting_writes_result_back(tmp_path):
    path = tmp_path / "candidates.xlsx"
    _make_excel(path)

    ExcelGreetingService.write_greeting_result(path, 2, "success", "已发送打招呼")
    workbook = load_workbook(path)
    sheet = workbook["候选人"]
    headers = {cell.value: index for index, cell in enumerate(sheet[1], start=1)}

    assert sheet.cell(row=2, column=headers["打招呼状态"]).value == "已发送"
    assert sheet.cell(row=2, column=headers["打招呼消息"]).value == "已发送打招呼"
    workbook.close()


def test_greeting_text_fallback_masks_salary():
    class FailingClient:
        def chat(self, prompt, system_message=""):
            raise RuntimeError("no api")

    service = GreetingTextGenerationService(FailingClient())
    text = service.generate(
        "销售总监",
        "深圳市南山区",
        "负责天然气设备客户开发和销售团队管理。薪资30-50万。",
        salary_range="30-50万",
    )

    assert text.startswith("您好，我是猎头顾问")
    assert "base深圳" in text
    assert "销售总监" in text
    assert "薪资可谈" in text
    assert "30-50" not in text
