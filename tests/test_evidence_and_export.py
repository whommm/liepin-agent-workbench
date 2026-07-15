from openpyxl import load_workbook
from zipfile import ZipFile

from liepin_agent.domain.models import CandidateDetail, CandidateSummary, MatchResult, SearchPlan
from liepin_agent.storage.sqlite_store import SQLiteStore
from liepin_agent.tools.exporter import ExportService
from liepin_agent.tools.real_matcher import RealMatchService


class BadJsonClient:
    def chat(self, prompt, system_message=""):
        return "not json"


def test_real_matcher_parse_failure_returns_review_package():
    matcher = RealMatchService(BadJsonClient())
    result = matcher.match_candidate(
        session_id="s1",
        round_id="r1",
        candidate_id="c1",
        resume_text="简历文本",
        criteria={"criteria_version_id": "v1", "keywords_text": "LNG"},
    )

    assert result.tier == ""
    assert result.status == "needs_review"
    assert result.criteria_version_id == "v1"
    assert result.confidence == "low"
    assert result.missing_or_unclear
    assert result.questions_to_verify


def test_evidence_labels_distinguish_exact_summary_and_legacy():
    assert ExportService._evidence_source_label(
        {"source_type": "direct", "grounding_status": "exact"}
    ) == "原文证据"
    assert ExportService._evidence_source_label(
        {"source_type": "direct", "grounding_status": "model_summary"}
    ) == "模型概括"
    assert ExportService._evidence_source_label(
        {"source_type": "direct"}
    ) == "匹配证据"
    assert ExportService._evidence_source_label(
        {"source_type": "inferred"}
    ) == "推断"


def test_export_contains_criteria_evidence_sources_and_metrics(tmp_path):
    store = SQLiteStore(str(tmp_path / "workbench.db"))
    session_id = store.create_session("销售总监", "天然气销售")
    criteria_id = store.create_criteria_version(
        session_id,
        "天然气\nLNG",
        "候选人需要具备天然气行业销售经验。",
        created_by="human",
    )
    store.confirm_criteria_version(criteria_id)
    plan = SearchPlan(
        query="天然气 销售",
        position_filter="销售",
        search_hypothesis_type="core_background",
        search_hypothesis_text="验证天然气销售背景",
    )
    round_id = store.create_round(session_id, 1, plan, criteria_id)
    candidate = CandidateSummary(
        session_id=session_id,
        round_id=round_id,
        profile_url="https://example.com/a",
        name="候选人",
        current_title="销售总监",
        current_company="天然气设备公司",
        card_decision="fetch",
        card_signals=["天然气"],
    )
    candidate_id = store.save_candidate_summary(candidate)
    store.save_candidate_source(
        candidate_id,
        session_id,
        round_id,
        criteria_id,
        plan,
        result_index=1,
        card_decision="fetch",
        card_signals=["天然气"],
    )
    store.save_candidate_detail(
        CandidateDetail(
            candidate_id=candidate_id,
            resume_text="天然气销售简历",
            capture_status="success",
            is_gold_collar=True,
        )
    )
    store.update_candidate_greeting_status(
        candidate_id, "success", message="已发送打招呼"
    )
    store.save_match_result(
        MatchResult(
            candidate_id=candidate_id,
            session_id=session_id,
            round_id=round_id,
            tier="B",
            summary="有天然气销售背景",
            risks="设备深度待确认",
            criteria_version_id=criteria_id,
            matched_evidence=[
                {
                    "criterion": "天然气",
                    "evidence": "负责天然气客户开发",
                    "strength": "strong",
                    "grounding_status": "exact",
                }
            ],
            missing_or_unclear=["压缩机经验未明"],
            questions_to_verify=["是否销售过压缩机设备？"],
            confidence="medium",
        )
    )
    rejected = CandidateSummary(
        session_id=session_id,
        round_id=round_id,
        profile_url="https://example.com/c",
        name="不合格候选人",
        current_title="销售经理",
        current_company="普通贸易公司",
        card_decision="fetch",
    )
    rejected_id = store.save_candidate_summary(rejected)
    store.save_match_result(
        MatchResult(
            candidate_id=rejected_id,
            session_id=session_id,
            round_id=round_id,
            tier="C",
            summary="天然气行业证据不足",
            criteria_version_id=criteria_id,
        )
    )
    unmatched = CandidateSummary(
        session_id=session_id,
        round_id=round_id,
        profile_url="https://example.com/pending",
        name="待复核候选人",
        current_title="销售负责人",
        current_company="能源公司",
        card_decision="maybe",
    )
    store.save_candidate_summary(unmatched)

    exporter = ExportService(store, tmp_path)
    path = exporter.export_session(session_id)
    workbook = load_workbook(path)

    assert "推荐总览" in workbook.sheetnames
    assert "合格A_B" in workbook.sheetnames
    assert "待复核_未匹配" in workbook.sheetnames
    assert "不合格C" in workbook.sheetnames
    assert "候选人" in workbook.sheetnames
    assert "寻访基准" in workbook.sheetnames
    assert "效率总结" in workbook.sheetnames
    assert "运行诊断" in workbook.sheetnames
    overview_headers = [cell.value for cell in workbook["推荐总览"][1]]
    assert "结论" in overview_headers
    assert "候选人档案" in overview_headers
    assert workbook["推荐总览"]["A2"].value == "可推荐"
    name_column = overview_headers.index("姓名") + 1
    assert workbook["合格A_B"].cell(row=2, column=name_column).value == "候选人"
    assert (
        workbook["待复核_未匹配"].cell(row=2, column=name_column).value
        == "待复核候选人"
    )
    assert (
        workbook["不合格C"].cell(row=2, column=name_column).value
        == "不合格候选人"
    )
    headers = [cell.value for cell in workbook["候选人"][1]]
    assert "命中证据" in headers
    assert "基准版本" in headers
    assert "简历链接" in headers
    assert "金领" in headers
    assert "打招呼状态" in headers
    evidence_column = headers.index("命中证据") + 1
    link_column = headers.index("简历链接") + 1
    gold_column = headers.index("金领") + 1
    greeting_column = headers.index("打招呼状态") + 1
    assert workbook["候选人"].cell(row=2, column=link_column).value == "https://example.com/a"
    assert workbook["候选人"].cell(row=2, column=link_column).hyperlink.target == "https://example.com/a"
    assert workbook["候选人"].cell(row=2, column=gold_column).value == "是"
    assert workbook["候选人"].cell(row=2, column=greeting_column).value == "已发送"
    assert workbook["候选人"].cell(
        row=2, column=evidence_column
    ).value.startswith("[原文证据]")
    assert workbook["寻访基准"]["B2"].value == "天然气\nLNG"
    diagnostics_values = [cell.value for cell in workbook["运行诊断"]["A"]]
    assert "待回写匹配数" in diagnostics_values
    assert "卡片判断" in diagnostics_values
    workbook.close()
    report_files = sorted(exporter.last_candidate_reports_dir.glob("*.docx"))
    assert len(report_files) == 3
    with ZipFile(report_files[0]) as archive:
        document_xml = archive.read("word/document.xml").decode("utf-8")
    assert "候选人" in document_xml
    assert "有天然气销售背景" in document_xml


def test_export_sanitizes_control_characters_in_session_title(tmp_path):
    store = SQLiteStore(str(tmp_path / "workbench.db"))
    session_id = store.create_session(
        "力传感器工程师\n\n1", "力传感器工程师"
    )

    exporter = ExportService(store, tmp_path)
    path = exporter.export_session(session_id)

    assert path.exists()
    assert path.name.startswith("力传感器工程师_1_")
    assert not any(ord(ch) < 32 for ch in path.name)
    assert exporter.last_candidate_reports_dir.is_dir()
    assert not any(
        ord(ch) < 32 for ch in exporter.last_candidate_reports_dir.name
    )


def test_export_repairs_legacy_card_field_misalignment(tmp_path):
    store = SQLiteStore(str(tmp_path / "workbench.db"))
    session_id = store.create_session("销售总监", "天然气销售")
    criteria_id = store.create_criteria_version(
        session_id,
        "天然气\n销售",
        "候选人需要具备天然气销售经验。",
        created_by="human",
    )
    store.confirm_criteria_version(criteria_id)
    plan = SearchPlan(query="天然气 销售", position_filter="销售")
    round_id = store.create_round(session_id, 1, plan, criteria_id)
    candidate = CandidateSummary(
        session_id=session_id,
        round_id=round_id,
        name="徐**",
        current_title="陕西澜山能源",
        current_company="有限责任公司 · 天然气销售",
        city="工作",
        work_years="16年",
        education="本科",
        summary_text="\n".join(
            [
                "徐**",
                "39岁",
                "工作16年",
                "本科",
                "西安-莲湖区",
                "求职期望：",
                "大连销售总监",
                "陕西澜山能源有限责任公司 · 天然气销售",
            ]
        ),
    )
    store.save_candidate_summary(candidate)

    path = ExportService(store, tmp_path).export_session(session_id)
    workbook = load_workbook(path)
    sheet = workbook["候选人"]
    headers = [cell.value for cell in sheet[1]]
    row = {header: sheet.cell(row=2, column=index + 1).value for index, header in enumerate(headers)}

    assert row["公司"] == "陕西澜山能源有限责任公司"
    assert row["职位"] == "天然气销售"
    assert row["城市"] == "西安-莲湖区"
    workbook.close()


def test_export_field_display_repairs_truncated_legacy_fields(tmp_path):
    store = SQLiteStore(str(tmp_path / "workbench.db"))
    exporter = ExportService(store, tmp_path)

    fields = exporter._display_candidate_fields(
        {
            "name": "王**",
            "current_title": "中燃宏大能源贸易",
            "current_company": "有限公司 · 区域销售总监",
            "city": "深圳",
            "work_years": "18年",
            "education": "本科",
            "summary_text": "",
        }
    )
    assert fields["current_company"] == "中燃宏大能源贸易有限公司"
    assert fields["current_title"] == "区域销售总监"

    fields = exporter._display_candidate_fields(
        {
            "name": "李**",
            "current_title": "SKF · 销售经理",
            "current_company": "",
            "city": "MBA/EMBA",
            "work_years": "16年",
            "education": "",
            "summary_text": "李**\n42岁\n工作16年\nMBA/EMBA\nZhuzhou",
        }
    )
    assert fields["current_company"] == "SKF"
    assert fields["current_title"] == "销售经理"
    assert fields["city"] == "Zhuzhou"
    assert fields["education"] == "MBA/EMBA"
def test_export_conclusion_treats_completed_d_as_rejected():
    assert ExportService._candidate_conclusion({"match_tier": "D"}) == "不推荐"
